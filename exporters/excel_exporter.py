"""ExcelExporter — exporteert een ReportPackage naar Excel (.xlsx).

JSON-sidecar formaat (naast .xlsx template, zelfde naam + .map.json):
{
  "metadata": {
    "project_name": {"sheet": "Voorblad", "cell": "B3"},
    "title":        {"sheet": "Voorblad", "cell": "B5"}
  },
  "sections": {
    "sheet_piling": "Damwand",
    "moment_max":   "Resultaten"
  }
}
"""

from __future__ import annotations
import json
from pathlib import Path

from reporting.models import ReportPackage, ReportSection


class ExcelExporter:
    """Schrijft een ReportPackage naar een .xlsx-bestand met openpyxl."""

    def export(self, package: ReportPackage, template_path: str | None,
               output_path: str) -> str | None:
        """Exporteer naar Excel.

        Returns:
            None bij succes, foutmelding (str) bij een uitzondering.
        """
        try:
            import openpyxl
        except ImportError:
            return 'openpyxl is niet geïnstalleerd. Voer uit: pip install openpyxl'

        try:
            mapping = self._load_mapping(template_path)

            if template_path and Path(template_path).exists():
                wb = openpyxl.load_workbook(template_path)
            else:
                wb = openpyxl.Workbook()
                if wb.active:
                    wb.active.title = 'Metadata'

            if mapping:
                self._write_with_mapping(wb, package, mapping)
            else:
                self._write_metadata(wb, package)
                all_sections = package.input_sections + package.result_sections
                selected_ids = {f'input_{i.source_ref}' for i in package.selected_items
                                if i.kind == 'invoer'} | \
                               {f'result_{i.source_ref}' for i in package.selected_items
                                if i.kind == 'resultaat'}
                for sec in all_sections:
                    # Als er geselecteerde items zijn, filter dan; anders schrijf alles
                    item_id_input = f'input_{sec.id}'
                    item_id_result = f'result_{sec.id}'
                    if selected_ids and item_id_input not in selected_ids \
                            and item_id_result not in selected_ids:
                        continue
                    self._write_section(wb, sec)

            wb.save(output_path)
            return None
        except Exception as exc:
            return str(exc)

    # ------------------------------------------------------------------
    # JSON-sidecar laden
    # ------------------------------------------------------------------

    def _load_mapping(self, template_path: str | None) -> dict | None:
        if not template_path:
            return None
        sidecar = Path(template_path).with_suffix('').with_suffix('.map.json')
        if not sidecar.exists():
            # probeer ook .xlsx.map.json
            sidecar2 = Path(str(template_path) + '.map.json')
            if not sidecar2.exists():
                return None
            sidecar = sidecar2
        try:
            return json.loads(sidecar.read_text(encoding='utf-8'))
        except Exception:
            return None

    # ------------------------------------------------------------------
    # Schrijven met sidecar-mapping
    # ------------------------------------------------------------------

    def _write_with_mapping(self, wb, package: ReportPackage, mapping: dict) -> None:
        """Vul template-cellen in via sidecar-mapping."""
        # Metadata-velden
        meta_map = mapping.get('metadata', {})
        md = package.metadata
        for attr, loc in meta_map.items():
            value = getattr(md, attr, '') or ''
            sheet_name = loc.get('sheet', '')
            cell = loc.get('cell', '')
            if sheet_name in wb.sheetnames and cell:
                wb[sheet_name][cell] = value

        # Secties: schrijf naar benoemde sheets als ze in mapping staan
        sec_map = mapping.get('sections', {})
        all_sections = package.input_sections + package.result_sections
        for sec in all_sections:
            sheet_name = sec_map.get(sec.id)
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                self._append_section_to_sheet(ws, sec)
            else:
                # Geen mapping → schrijf als nieuw werkblad
                self._write_section(wb, sec)

    def _append_section_to_sheet(self, ws, section: ReportSection) -> None:
        """Voeg sectie-data toe aan een bestaand werkblad (achter bestaande data)."""
        start_row = ws.max_row + 2
        row = start_row
        for f in section.fields:
            ws.cell(row=row, column=1, value=f.label)
            val = f'{f.value} {f.unit}'.strip() if f.unit else f.value
            ws.cell(row=row, column=2, value=val)
            row += 1
        for table in section.tables:
            row += 1
            ws.cell(row=row, column=1, value=table.title)
            row += 1
            for col, header in enumerate(table.columns, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            for data_row in table.rows:
                for col, cell in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col, value=cell)
                row += 1
        for tb in section.text_blocks:
            row += 1
            ws.cell(row=row, column=1, value=tb.effective_text)
            row += 1

    # ------------------------------------------------------------------
    # Standaard schrijven (geen mapping)
    # ------------------------------------------------------------------

    def _write_metadata(self, wb, package: ReportPackage) -> None:
        ws = (wb['Metadata'] if 'Metadata' in wb.sheetnames
              else wb.create_sheet('Metadata'))
        md = package.metadata
        rows = [
            ('Projectnaam',    md.project_name),
            ('Ordernummer',    md.order_number),
            ('Locatie',        md.location),
            ('Fase',           md.phase),
            ('Opdrachtgever',  md.client),
            ('Titel',          md.title),
            ('Revisie',        md.revision),
            ('Auteur',         md.author),
            ('Datum',          md.date),
        ]
        for r, (label, value) in enumerate(rows, start=1):
            ws.cell(row=r, column=1, value=label)
            ws.cell(row=r, column=2, value=value)

    def _write_section(self, wb, section: ReportSection) -> None:
        sheet_name = section.title[:31]
        if sheet_name in wb.sheetnames:
            sheet_name = sheet_name[:28] + '...'
        ws = wb.create_sheet(sheet_name)

        row = 1
        for f in section.fields:
            ws.cell(row=row, column=1, value=f.label)
            val = f'{f.value} {f.unit}'.strip() if f.unit else f.value
            ws.cell(row=row, column=2, value=val)
            row += 1

        for table in section.tables:
            if row > 1:
                row += 1
            ws.cell(row=row, column=1, value=table.title)
            row += 1
            for col, header in enumerate(table.columns, start=1):
                ws.cell(row=row, column=col, value=header)
            row += 1
            for data_row in table.rows:
                for col, cell in enumerate(data_row, start=1):
                    ws.cell(row=row, column=col, value=cell)
                row += 1

        for tb in section.text_blocks:
            if row > 1:
                row += 1
            ws.cell(row=row, column=1, value='Beschrijving')
            ws.cell(row=row, column=2, value=tb.effective_text)
            row += 1
